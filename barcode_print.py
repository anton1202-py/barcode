from collections import Counter
import glob
import os
from pathlib import Path
import time
from contextlib import closing
import io


import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font
import shutil
import pandas as pd
from PyQt6.QtCore import QSettings, Qt
from PyQt6.QtWidgets import (QApplication, QCheckBox, QFileDialog,
                             QGridLayout, QHBoxLayout, QLabel,
                             QLineEdit, QListWidget,QPushButton,
                             QRadioButton, QWidget, QVBoxLayout)

from helper_func import (print_barcode_to_pdf, 
                         qrcode_print_to_file_main, design_barcodes)


class Interface(QWidget):
    def __init__(self):
        super().__init__()
        from main_file import dbx_db, version

        self.dbx = dbx_db
        self.settings = QSettings('settings.ini', QSettings.Format.IniFormat)
        self.settings.setFallbacksEnabled(False)

        self.setStyleSheet("background-color: #2a2c2e; font: 12pt; color: #f7f7f7;")
        self.resize(350, 400)
        self.setWindowTitle(f"Иннотрейд {version}")

        self.setAcceptDrops(True)
        self.files_list = []
        self.xls_files_list = []
        self.list_files = QListWidget()
        self.list_files.setStyleSheet("font: 10pt; color: #f7f7f7;")
        self.label_total_files = QLabel()

        layout = QVBoxLayout()
        layout.addWidget(QLabel('Перетащите файл с количеством в это окно:'))
        layout.addWidget(self.list_files)
        layout.addWidget(self.label_total_files)
        layout.setContentsMargins(50, 50, 50, 50)
        self.setLayout(layout)

        self.label = QLabel("Выберите юр. лицо для этикетки")
        self.label.setStyleSheet(
            '''QLabel {
                font: 14pt;
                color: #f7f7f7;
                margin-top: 40px;
                }''')
        layout.addWidget(self.label)

        self.radiobtn1 = QRadioButton("ООО Иннотрейд", self)
        self.radiobtn1.setStyleSheet(
            '''QRadioButton {font: 12pt; color: #f7f7f7;}''')
        self.radiobtn1.toggled.connect(self.showDetails)
        layout.addWidget(self.radiobtn1)

        self.radiobtn2 = QRadioButton("ИП Караваев", self)
        self.radiobtn2.setStyleSheet(
            '''QRadioButton {font: 12pt; color: #f7f7f7;}''')
        self.radiobtn2.toggled.connect(self.showDetails)
        layout.addWidget(self.radiobtn2)

        self.label1 = QLabel("", self)
        layout.addWidget(self.label1)

        layout_2 = QHBoxLayout()
        layout.addLayout(layout_2)

        self.check1 = QCheckBox("Добавить QR коды")
        self.check1.setStyleSheet(
            '''QCheckBox {font: 12pt; color: #f7f7f7; margin-top: 20px;}''')
        self.check1.stateChanged.connect(self.checkbox_status)
        layout_2.addWidget(self.check1)

        self.btn3 = QPushButton("Выбрать файл с QR кодами")
        self.btn3.setStyleSheet(
            '''QPushButton {
                border: 2px solid #f74a00;
                font: 12pt; color: #f7f7f7;
                border-radius: 8px;
                width: 250px; height: 40px;
                margin-top: 20px;
                }
                QPushButton::hover {
                background-color: #f74a00;}
                QPushButton:disabled {
                        background-color: #575757;
                        border: 0px;}
                QPushButton::pressed {
                        background-color: #ff1e00;}
                ''')
        self.btn3.setEnabled(False)
        self.btn3.pressed.connect(self.chosePdflFile)
        layout_2.addWidget(self.btn3)

        self.check_pdf_assembly = QCheckBox("Формировать остатки FBS")
        self.check_pdf_assembly.setStyleSheet(
            '''QCheckBox {font: 12pt; color: #f7f7f7; margin-top: 20px;}
               QCheckBox:disabled {
                        color: #8c8c8c;}''')
        self.check_pdf_assembly.stateChanged.connect(self.checkbox_status)
        layout.addWidget(self.check_pdf_assembly)
        self.check_pdf_assembly.setEnabled(False)

        self.btn = QPushButton("Сформировать итоговый PDF")
        self.btn.setStyleSheet(
            '''QPushButton {
                background-color: #f74a00;
                font: 12pt; color: #f7f7f7;
                border-radius: 8px;
                height: 40px;
                margin-top: 30px;
                }
                QPushButton:disabled {
                        background-color: #575757;
                        border: 0px;}
                QPushButton::pressed {
                        background-color: #ff1e00;}
                ''')
        self.btn.pressed.connect(self.check_push_radiobutton)
        self.btn.setEnabled(False)
        layout.addWidget(self.btn)


    def _update_states(self):
        self.label_total_files.setText('Загружено файлов: {}'.format(self.list_files.count()))
        if len(self.list_files) == 0:
            self.btn.setEnabled(False)
        else:
            self.btn.setEnabled(True)

    def dragEnterEvent(self, event):
        # Тут выполняются проверки и дается (или нет) разрешение на Drop
        mime = event.mimeData()
        # Если перемещаются ссылки
        if mime.hasUrls():
            # Разрешаем
            event.acceptProposedAction()
        
    def dropEvent(self, event):
        # Обработка события Drop
        for url in event.mimeData().urls():
            file_name = url.toLocalFile()
            self.list_files.addItem(file_name)
            self.files_list.append(file_name)
        self._update_states()
        return super().dropEvent(event)

    def func_returne_list(self):
        return self.files_list
    
    def remove(self):
        current_row = self.list_files.currentRow()
        if current_row >= 0:
            current_item = self.list_files.takeItem(current_row)
            del current_item

    def keyPressEvent(self, e):
        """Удаляет строки из списка файлов при нажатии на delete"""
        if e.key() == Qt.Key.Key_Delete:
            self.remove()
        if self.list_files.count() == 0:
            self.btn.setEnabled(False)
        else:
            self.btn.setEnabled(True)

    def checkbox_status(self):
        """Проверяет статус чекбокса"""
        if self.check1.isChecked():
            self.btn3.setEnabled(True)
            self.check_pdf_assembly.setEnabled(True)
        else:
            self.btn3.setEnabled(False)
            self.check_pdf_assembly.setEnabled(False)

    def choseExcelFile(self):
        """Выбирает Excel файл, где указано кол-во этикеток"""
        filename = self.files_list[0]
        if filename:
            path = Path(filename)
        global filename_xls
        filename_xls = str(filename)
        global file_name_dir
        file_name_dir = path.parent
        global filename_for_print
        filename_for_print = os.path.basename(path).split('.')[0]
        return (filename_xls)

    def chosePdflFile(self):
        """Выбирает файл с QR кодами"""
        layout1 = QGridLayout()
        self.setLayout(layout1)
        self.filename_edit = QLineEdit()
        layout1.addWidget(QLabel('label_2'), 1, 1)
        layout1.addWidget(self.filename_edit, 0, 1)
        self.show()
        filename, ok = QFileDialog.getOpenFileName(
            self,
            "Выберите PDF файл с QR кодами",
            self.settings.value('Lastfile'),
            "Files (*.pdf)"
        )
        if filename:
            path = Path(filename)
            self.filename_edit.setText(str(path))
        global filename_for_qrcodes
        global filename_pdf
        filename_for_qrcodes = os.path.basename(path).split('.')[0]
        filename_pdf = str(filename)

    def stream_dropbox_file(self, path):
        _,res=self.dbx.files_download(path)
        with closing(res) as result:
            byte_data=result.content
            return io.BytesIO(byte_data)

    def showDetails(self):
        """Логика работы радиокнопок"""
        global main_file
        if self.radiobtn1.isChecked():
            path = '/DATABASE/Ночники ООО.xlsx'
            main_file = self.stream_dropbox_file(path)
            path_for_ticket = '/DATABASE/helper_files/Печать Иннотрейд.xlsx'
            return self.stream_dropbox_file(path_for_ticket)
        elif self.radiobtn2.isChecked():
            path = '/DATABASE/Ночники ИП.xlsx'
            main_file = self.stream_dropbox_file(path)
            path_for_ticket = '/DATABASE/helper_files/Печать Караваев.xlsx'
            return self.stream_dropbox_file(path_for_ticket)
        else:
            print('Выберите юр. лицо!')

    def check_push_radiobutton(self):
        """Обрабатывает ошибку не выбора радиокнопок"""
        # checking if it is checked
        if self.radiobtn1.isChecked() or self.radiobtn2.isChecked():
            self.check_excel_file()
        else:
            # changing text of label
            self.label1.setText(
                '<h3 style="color: rgb(250, 55, 55);">Выберете юр. лицо!</h3>'
                )

    def check_excel_file(self):
        self.choseExcelFile()
        if len(self.files_list) == 0:
            self.label1.setText(
                '''<h3 style="color: rgb(250, 55, 55);">
                Выберете excel файл с количеством</h3>''')
        else:
            self.check_address()

    def check_address(self):
        """Проверяет, что выбран файл с нужным юр. лцом"""
        # Читаем в файле excel нужные столбцы
        amount_file_location = filename_xls
        excel_data1 = pd.read_excel(amount_file_location)
        excel_data2 = openpyxl.load_workbook(self.showDetails())
        sheet = excel_data2.active
        data1 = pd.DataFrame(excel_data1, columns=['Артикул продавца', 'Количество'])

        # Список только нужных наименований
        name_for_print = data1['Артикул продавца'].to_list()
        list_first_letter = []

        for article in name_for_print:
            list_first_letter.append(article[0])

        if (sheet["A4"].value == 'ООО “Иннотрейд”, Адрес: '
                and ('L' not in list_first_letter and
                     'V' not in list_first_letter)
            ) or (sheet["A4"].value == 'ИП Караваев Е.Г.'
                  and ('S' not in list_first_letter and 'j' not in
                  list_first_letter and 'N' not in list_first_letter)):
            self.print_barcode_to_pdf()
        else:
            self.label1.setText(
                '''<h3 style="color: rgb(250, 55, 55);">
                Выберете другое юр. лицо!</h3>''')


class BarcodePrint(Interface):
    def __init__(self):
        super().__init__()

    def qrcode_print_to_file(self):
        """Создает QR коды в необходимом формате"""
        global pdf_filenames_qrcode
        pdf_filenames_qrcode = qrcode_print_to_file_main(
            filename_pdf, filename_for_qrcodes)

    def create_list_barcode(self):
        """Формирует список штрихкодов для печати"""
        if self.check1.isChecked():
            self.qrcode_print_to_file()
        # Читаем в файле excel нужные столбцы
        main_file_location = main_file
        amount_file_location = filename_xls

        excel_data = pd.read_excel(main_file_location)
        excel_data1 = pd.read_excel(amount_file_location)
        excel_data2 = openpyxl.load_workbook(self.showDetails())

        sheet = excel_data2.active
        data = pd.DataFrame(
            excel_data, columns=['Артикул продавца', 'Баркод товара', 'Наименование'])
        data1 = pd.DataFrame(excel_data1, columns=['Артикул продавца', 'Количество'])

        lenght_name_for_print_raw = len(data1['Артикул продавца'].to_list())
        # Список только нужных наименований
        name_for_print_raw = data1['Артикул продавца'].to_list()
        name_for_print_raw = [item.capitalize() for item in name_for_print_raw]
        amount_for_print_raw = data1['Количество'].to_list()
        amount_for_print_raw2 = list()

        for element in amount_for_print_raw:
            if str(element) == "nan":
                amount_for_print_raw2.append(1)
            else: 
                amount_for_print_raw2.append(element)

        barcodelist_list = data['Баркод товара'].to_list()
        article_list = data['Артикул продавца'].to_list()
        article_list = [item.capitalize() for item in article_list]
        name_list = data['Наименование'].to_list()

        design_barcodes(name_for_print_raw, article_list,
                        amount_file_location, main_file_location,
                        barcodelist_list, name_list, sheet)

        pdf_filenames = glob.glob('cache_dir/*.pdf')
        new_list_raw = []
        for i in range(lenght_name_for_print_raw):
            for j in pdf_filenames:
                if str(name_for_print_raw[i]) == str(Path(j).stem):
                    while amount_for_print_raw2[i] > 0:
                        new_list_raw.append(j)
                        amount_for_print_raw2[i] -= 1

        # Создаем сводный файл
        wb = openpyxl.Workbook()
        shs11 = wb.create_sheet(title = 'pivot_list', index = 0)
        sheet = wb['pivot_list']
        sheet['A1'] = 'Артикул продавца'
        sheet['B1'] = 'Количество'
        list_for_excel = []
        for i in new_list_raw:
            s = i.split('\\')
            m = s[1].split('.')
            list_for_excel.append(m[0])
        counter = Counter(list_for_excel)
        i_1 = 1
        for r in counter:
            shs11.cell(row=i_1+1,column=1).value=r
            shs11.cell(row=i_1+1,column=2).value=counter[r]
            i_1 += 1
        wb.save(f'cache_dir/raw_excel.xlsx')
        # Читаем созданные сводный файл
        excel_data3 = pd.read_excel('cache_dir/raw_excel.xlsx')
        data3 = pd.DataFrame(excel_data3, columns=['Артикул продавца', 'Количество'])
        name_for_print = data3['Артикул продавца'].to_list()
        name_for_print = [item.capitalize() for item in name_for_print]
        amount_for_print = data3['Количество'].to_list()
        lenght_name_for_print = len(data3['Артикул продавца'].to_list())

        global new_list
        new_list = list()
        for i in range(lenght_name_for_print):
            pdf_filenames.sort()
            for j in pdf_filenames:
                if str(name_for_print[i]) == str(Path(j).stem):
                    while amount_for_print[i] > 0:
                        new_list.append(j)
                        amount_for_print[i] -= 1
        # Если чек-бокс нажат, создаю сводный excel файл для производства
        if self.check1.isChecked():
            wb = openpyxl.Workbook()
            shs11 = wb.create_sheet(title = 'pivot_list', index = 0)
            sheet = wb['pivot_list']
            sheet['A1'] = 'Артикул'
            sheet['B1'] = 'Производство'
            sheet['C1'] = 'FBS'
            sheet['D1'] = 'Сборщикам'

            list_for_excel = []
            new_list_raw.sort()
            for i in new_list_raw:
                s = i.split('\\')
                m = s[1].split('.')
                list_for_excel.append(m[0])
            counter = Counter(list_for_excel)
            i_1 = 1
            PROD_DETAIL_CONST = 4
            for r in counter:
                # Заполняет столбец ['A1'] = 'Артикул продавца'
                shs11.cell(row=i_1+1,column=1).value=r
                # Заполняет столбец ['B1'] = 'FBS'
                shs11.cell(row=i_1+1,column=3).value=int(counter[r])
                # Заполняет столбец ['C1'] = 'Произвдство'
                if counter[r] == 1:
                    shs11.cell(row=i_1+1,column=2).value = int(PROD_DETAIL_CONST)
                elif 2<= counter[r] <= PROD_DETAIL_CONST-1:
                    shs11.cell(row=i_1+1,column=2).value= int(2 * PROD_DETAIL_CONST)
                elif PROD_DETAIL_CONST<= counter[r] <= 2 * PROD_DETAIL_CONST - 1:
                    shs11.cell(row=i_1+1,column=2).value= int(3 * PROD_DETAIL_CONST)
                else:
                    shs11.cell(row=i_1+1,column=2).value= ' '
                i_1 += 1

            wb.save(f'{file_name_dir}/Производство {filename_for_print} '
                    f'{time.strftime("%Y-%m-%d %H-%M")}.xlsx')

            file_path = (f'{file_name_dir}/Производство {filename_for_print} '
                f'{time.strftime("%Y-%m-%d %H-%M")}.xlsx')
            w_b = load_workbook(file_path)
            source_page = w_b.active

            name_article = source_page['A']
            amount_production = source_page['B']
            amount_fbs = source_page['C']

            # Заполняет столбец ['D1'] = 'Сборщикам'
            for r in range(1, len(name_article)):
                if amount_production[r].value != ' ':
                    source_page.cell(row=r+1,column=4).value = int(amount_production[r].value) - int(amount_fbs[r].value)
            w_b.save(file_path)

            w_b2 = load_workbook(file_path)
            source_page2 = w_b2.active
            al = Alignment(horizontal="center", vertical="center")
            # Задаем толщину и цвет обводки ячейки
            font_bold = Font(bold=True)
            thin = Side(border_style="thin", color="000000")
            for i in range(len(amount_production)):
                for c in source_page2[f'A{i+1}:D{i+1}']:
                        c[0].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                        c[1].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                        c[2].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                        c[3].border = Border(top=thin, right=thin, left=thin, bottom = thin)

                        c[0].alignment = al
                        c[1].alignment = al
                        c[2].alignment = al
                        c[3].alignment = al

                        c[0].font = font_bold
                        c[1].font = font_bold

            source_page2['C1'].font = font_bold
            source_page2['D1'].font = font_bold
            
            source_page2.column_dimensions['A'].width = 18
            source_page2.column_dimensions['B'].width = 18
            source_page2.column_dimensions['C'].width = 18
            source_page2.column_dimensions['D'].width = 18
            w_b2.save(file_path)
            #self.create_excel_file_for_new_system()
            
            if not self.check_pdf_assembly.isChecked():
                w_b = load_workbook(file_path)
                source_page = w_b.active
                source_page.delete_cols(2, 1)
                source_page.delete_cols(4, 1)
                w_b.save(file_path)
            elif self.check_pdf_assembly.isChecked():
                self.create_pdf_file_ticket_for_complect()
        # Сортирую данные в папке по возрастанию чисел
        new_list.sort()
        try:
            if pdf_filenames_qrcode:
                for j in pdf_filenames_qrcode:
                    new_list.append(j)
        except:
            print(f'')
        return (new_list)

    
    def create_pdf_file_ticket_for_complect(self):
        """Создает PDF-файл для формирования остатки FBS"""
        first_file = pd.read_excel(f'{file_name_dir}/Производство {filename_for_print} '
                f'{time.strftime("%Y-%m-%d %H-%M")}.xlsx')
        
        first_file_data = pd.DataFrame(first_file, columns=['Артикул', 'Сборщикам'])
        
        name_ticket_for_print = first_file_data['Артикул'].to_list()
        amount_first_file = first_file_data['Сборщикам'].to_list()
    
        amount_for_print = []
        for i in range(len(name_ticket_for_print)):
            if str(amount_first_file[i]) == 'nan':
                amount_for_print.append(1)
            else:
                quantity_for_print = amount_first_file[i] + 1
                amount_for_print.append(quantity_for_print)

        pdf_filenames = glob.glob('cache_dir/*.pdf')
        list_pdf_file_ticket_for_complect = []
        for i in range(len(name_ticket_for_print)):
            for j in pdf_filenames:
                if str(name_ticket_for_print[i]) == str(Path(j).stem):
                    while amount_for_print[i] > 0:
                        list_pdf_file_ticket_for_complect.append(j)
                        amount_for_print[i] -= 1
        file_name = (f'{file_name_dir}/Наклейки для комплектовщиков '
                         f'{time.strftime("%Y-%m-%d %H-%M")}.pdf')
        print_barcode_to_pdf(list_pdf_file_ticket_for_complect, file_name)

    def print_barcode_to_pdf(self):
        """Создает pdf файл для печати"""
        file_name = (f'{file_name_dir}/{filename_for_print} '
                         f'{time.strftime("%Y-%m-%d %H-%M")}.pdf')
        print_barcode_to_pdf(self.create_list_barcode(), file_name)
        
        # Очистка переменных перед новым запуском
        new_list.clear()
        self.files_list.clear()
        self.label1.clear()
        globals().pop('filename_xls')

        # Очистка кеша
        dir = 'cache_dir/'
        filelist = glob.glob(os.path.join(dir, "*"))
        for f in filelist:
            try:
                os.remove(f)
            except Exception:
                print(' ')
        
        dir = 'cache_dir_2/'
        filelist = glob.glob(os.path.join(dir, "*"))
        for f in filelist:
            try:
                os.remove(f)
            except Exception:
                print(' ')

        try:
            dir = f'cache_dir_3/{pdf_filenames_qrcode}'
            filelist = glob.glob(os.path.join(dir, "*"))
            for f in filelist:
                try:
                    os.remove(f)
                except Exception:
                    print('')
            print('Done')
        except Exception:
                print('')

        folder = 'cache_dir_3/'
        for filename in glob.glob(os.path.join(folder, "*")):
            file_path = os.path.join(folder, filename)
            try:
                if os.path.isfile(filename) or os.path.islink(filename):
                    os.unlink(filename)
                elif os.path.isdir(filename):
                    shutil.rmtree(filename)
            except Exception as e:
                print('')

        print('Done!')


    dir = 'cache_dir_2/'
    filelist = glob.glob(os.path.join(dir, "*"))
    for f in filelist:
        os.remove(f)


    folder = 'cache_dir_3/'
    for filename in glob.glob(os.path.join(folder, "*")):
        file_path = os.path.join(folder, filename)
        try:
            if os.path.isfile(filename) or os.path.islink(filename):
                os.unlink(filename)
            elif os.path.isdir(filename):
                shutil.rmtree(filename)
        except Exception as e:
            print('')
    
    dir = 'cache_dir_3/'
    filelist = glob.glob(os.path.join(dir, "*"))
    for f in filelist:
        os.remove(f)

    # Удаление кеша из папки с кешем
    dir = 'cache_dir/'
    filelist = glob.glob(os.path.join(dir, "*"))
    for f in filelist:
        os.remove(f)
    

if __name__ == '__main__':
    import sys
    app = QApplication(sys.argv)
    w = BarcodePrint()
    w.show()
    sys.exit(app.exec())
