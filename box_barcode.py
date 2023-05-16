from barcode import Code128
from barcode.writer import ImageWriter
import glob
import img2pdf
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import os
from contextlib import closing
import io
import pandas as pd
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
from PyQt6.QtCore import QSettings
from PyQt6.QtWidgets import (QApplication, QFileDialog,
                             QGridLayout,  QLabel, 
                             QLineEdit, QPushButton, QRadioButton,
                             QWidget, QVBoxLayout)
import PySimpleGUI as sg
import pythoncom
import time
from win32com.client import DispatchEx

from helper_func import print_barcode_to_pdf2


class BarcodeBoxPrintToPdf(QWidget):
    def __init__(self):
        super().__init__()
        from main_file import dbx_db, version
        
        self.dbx = dbx_db
        self.settings = QSettings('settings.ini', QSettings.Format.IniFormat)
        self.settings.setFallbacksEnabled(False)
        self.setStyleSheet("background-color: #2a2c2e;")
        self.resize(350, 300)
        self.setWindowTitle(f"Иннотрейд. Печать этикеток для коробок. {version}")

        layout = QVBoxLayout()
        layout.setContentsMargins(50, 50, 50, 50)
        self.setLayout(layout)

        btn2 = QPushButton("Выбрать файл с заказом")
        btn2.setStyleSheet(
            '''QPushButton {
                border: 2px solid #f74a00;
                font: 12pt; color: #f7f7f7;
                border-radius: 8px;
                height: 40px;
                }
                QPushButton::hover {
                background-color: #f74a00;
                font: 12pt; color: #f7f7f7;
                border-radius: 8px;
                height: 40px;
                }
                QPushButton::pressed {
                        background-color: #ff1e00;}
                ''')
        btn2.pressed.connect(self.chose_ExcelFile_product)
        layout.addWidget(btn2)

        self.label_path_product = QLabel("", self)
        self.label_path_product.setStyleSheet(
        '''QLabel {
                font: 12pt; color: #f7f7f7;
                margin-top: 0px;
                }
                ''')
        layout.addWidget(self.label_path_product)

        btn4 = QPushButton("Выбрать шаблон ВБ с этикетками")
        btn4.setStyleSheet(
            '''QPushButton {
                border: 2px solid #f74a00;
                font: 12pt; color: #f7f7f7;
                border-radius: 8px;
                height: 40px;
                }
                QPushButton::hover {
                background-color: #f74a00;
                font: 12pt; color: #f7f7f7;
                border-radius: 8px;
                height: 40px;
                }
                QPushButton::pressed {
                        background-color: #ff1e00;}
                ''')
        btn4.pressed.connect(self.chose_ExcelFile_box)
        layout.addWidget(btn4)

        self.label_path_box = QLabel("", self)
        self.label_path_box.setStyleSheet(
        '''QLabel {
                font: 12pt; color: #f7f7f7;
                margin-top: 0px;
                }
                ''')
        layout.addWidget(self.label_path_box)

        self.check_error = QLabel("", self)
        self.check_error.setStyleSheet(
        '''QLabel {
                font: 12pt;
                color: #ff5c5c;
                margin-top: 10px;
                }
                ''')
        layout.addWidget(self.check_error)
        self.label = QLabel("Выберите юр. лицо для этикетки")
        self.label.setStyleSheet(
            '''QLabel {
                font: 14pt;
                color: #f7f7f7;
                margin-top: 0px;
                }''')
        layout.addWidget(self.label)

        self.radiobtn1 = QRadioButton("ООО Иннотрейд", self)
        self.radiobtn1.setStyleSheet(
            '''QRadioButton {font: 12pt; color: #f7f7f7;}''')
        self.radiobtn1.toggled.connect(self.showDetails)
        layout.addWidget(self.radiobtn1)

        self.radiobtn2 = QRadioButton("ИП Караваев", self)
        self.radiobtn2.setStyleSheet(
            '''QRadioButton {font: 12pt; color: #f7f7f7;}''')
        self.radiobtn2.toggled.connect(self.showDetails)
        layout.addWidget(self.radiobtn2)

        btn = QPushButton("Сформировать итоговый PDF и файл для сборки")
        btn.setStyleSheet(
            '''QPushButton {
                background-color: #f74a00;
                font: 12pt; color: #f7f7f7;
                border-radius: 8px;
                height: 40px;
                width: 400px;
                margin-top: 30px;
                }
                QPushButton::pressed {
                        background-color: #ff1e00;}
                ''')
        #btn.setEnabled(False)
        btn.pressed.connect(self.check_address)
        layout.addWidget(btn)


    def showDetails(self):
        """Логика работы радиокнопок"""
        global database_file
        if self.radiobtn1.isChecked():
            database_file = '/DATABASE/Ночники ООО.xlsx'
        elif self.radiobtn2.isChecked():
            database_file = '/DATABASE/Ночники ИП.xlsx'
        else:
            print('Выберите юр. лицо!')


    def chose_ExcelFile_product(self):
        """Выбирает Excel файл с заказом"""
        layout1 = QGridLayout()
        self.setLayout(layout1)
        self.filename_edit = QLineEdit()
        layout1.addWidget(QLabel('label_2'), 1, 1)
        layout1.addWidget(self.filename_edit, 0, 1)
        self.show()
        filename, ok = QFileDialog.getOpenFileName(
            self,
            "Выберите Excel файл с заказом",
            self.settings.value('Lastfile'),
            "Files (*.xls *.xlsx)"
        )
        self.settings.setValue('Lastfile', filename)
        if filename:
            path = Path(filename)
            self.filename_edit.setText(str(path))
        self.label_path_product.setText(f'{path}')
        global filename_xls_product
        filename_xls_product = str(filename)
        global file_name_dir_product
        file_name_dir_product = path.parent
        global filename_for_print
        filename_for_print = os.path.basename(path).split('.')[0]
        return (filename_xls_product)


    def chose_ExcelFile_box(self):
        """Выбирает Excel файл - шаблон этикеток ВБ"""
        layout1 = QGridLayout()
        self.setLayout(layout1)
        self.filename_edit = QLineEdit()
        layout1.addWidget(QLabel('label_2'), 1, 1)
        layout1.addWidget(self.filename_edit, 0, 1)
        self.show()
        filename, ok = QFileDialog.getOpenFileName(
            self,
            "Выберите Excel шаблон этикеток ВБ",
            self.settings.value('Lastfile'),
            "Files (*.xls *.xlsx)"
        )
        self.settings.setValue('Lastfile', filename)
        if filename:
            path = Path(filename)
            self.filename_edit.setText(str(path))
        self.label_path_box.setText(f'{path}')
        global filename_xls_box
        filename_xls_box = str(filename)
        global file_name_dir_box
        file_name_dir_box = path.parent

        global filename_for_print
        filename_for_print = os.path.basename(path).split('.')[0]
        return (filename_xls_box)


    def check_address(self):
        """Проверяет, что файлы для обработки выбраны"""
        if not self.label_path_product.text() and not self.label_path_box.text():
            self.check_error.setText('выберете файлы со штрихкодами коробок и товаров')
            
        elif not self.label_path_product.text():
            self.check_error.setText('выберете файл со штрихкодами коробок')
        elif not self.label_path_box.text():
            self.check_error.setText('выберете файл со штрихкодами товаров')
        elif not self.radiobtn1.isChecked() and not self.radiobtn2.isChecked():
            self.check_error.setText('выберете юр. лицо')
        else:
            self.read_xls_file()
            self.create_xls_file()
            self.create_xls_file_for_WB()
            self.style_xls_file()
            self.export_to_pdf()
            self.print_barcode_to_pdf()

    def stream_dropbox_file(self, path):
        _,res=self.dbx.files_download(path)
        with closing(res) as result:
            byte_data=result.content
            return io.BytesIO(byte_data)

    def read_xls_file(self):
        """Функция читает исходные данные и формирует новые"""
        # Читаем первый входящий файл
        excel_data = pd.read_excel(filename_xls_product)
        data = pd.DataFrame(excel_data, columns = ['баркод',
                                                   'количество коробок',
                                                   'количество товара'])
        barcode_list = data['баркод'].to_list()
        amount_list = data['количество коробок'].to_list()
        amount_product_list = data['количество товара'].to_list()

        # Читаем второй входящий файл
        excel_data_1 = pd.read_excel(filename_xls_box)
        data_1 = pd.DataFrame(excel_data_1, columns = ['шк короба'])
        global barcode_boxes_list
        barcode_boxes_list = data_1['шк короба'].to_list()

        # Создаем словарь {'порядковый номер': 'количество товара в коробке'}
        dict_numb_barcode = {} # Словарь {'порядковый номер': 'баркод товара'}
        for i in range(len(barcode_list)):
            dict_numb_barcode[i]= barcode_list[i]

        dict_numb_amount_box = {} # Словарь {'порядковый номер': 'количество коробок'}
        for i in range(len(amount_list)):
            dict_numb_amount_box[i]= amount_list[i]

        dict_numb_amount_product_in_box = {} # Словарь {'порядковый номер': 'количество коробок'}
        for i in range(len(amount_product_list)):
            dict_numb_amount_product_in_box[i]= amount_product_list[i]

        sorted_values_for_dict_numb_barcode = sorted(
            dict_numb_barcode.values(), key=lambda x: str(x)[-4:]) # Sort the values
        sorted_dict_numb_barcode = {}


        for i in sorted_values_for_dict_numb_barcode:
            for k in dict_numb_barcode.keys():
                if dict_numb_barcode[k] == i:
                    sorted_dict_numb_barcode[k] = dict_numb_barcode[k]

        list_vizhual_code_for_people = []
        list_vizhual_code_for_ticket = []
        for data_m in sorted_dict_numb_barcode.values():
            list_vizhual_code_for_people.append(str(data_m)[-4:])

        path = database_file
        main_file = self.stream_dropbox_file(path)
        excel_data_3 = pd.read_excel(main_file)
        data_3 = pd.DataFrame(excel_data_3, columns = ['Артикул продавца',
                                                   'Баркод товара'])
        barcode_main_list = data_3['Баркод товара'].to_list()
        article_main_list = data_3['Артикул продавца'].to_list()


        sorted_dict_numb_barcode_values = []
        for i in sorted_dict_numb_barcode.values():
            sorted_dict_numb_barcode_values.append(i)

        for i in range(len(sorted_dict_numb_barcode_values)):
            if i > 0 and i < len(sorted_dict_numb_barcode_values) - 1:
                if str(sorted_dict_numb_barcode_values[i])[-4:] == str(
                    sorted_dict_numb_barcode_values[i-1])[-4:] or str(
                    sorted_dict_numb_barcode_values[i])[-4:] == str(
                    sorted_dict_numb_barcode_values[i+1])[-4:]:
                    ind = barcode_main_list.index(str(sorted_dict_numb_barcode_values[i]))
                    list_vizhual_code_for_ticket.append(article_main_list[ind].capitalize())
                else:
                    list_vizhual_code_for_ticket.append(' ')
            elif i == 0 :
                if str(sorted_dict_numb_barcode_values[i])[-4:] == str(
                    sorted_dict_numb_barcode_values[i+1])[-4:]:
                    ind = barcode_main_list.index(str(
                        sorted_dict_numb_barcode_values[i]))
                    list_vizhual_code_for_ticket.append(article_main_list[ind].capitalize())
                else:
                    list_vizhual_code_for_ticket.append(' ')
            elif i >= len(sorted_dict_numb_barcode_values) - 1 :
                if str(sorted_dict_numb_barcode_values[i])[-4:] == str(
                    sorted_dict_numb_barcode_values[i-1])[-4:]:
                    ind = barcode_main_list.index(str(sorted_dict_numb_barcode_values[i]))
                    list_vizhual_code_for_ticket.append(article_main_list[ind].capitalize())
                else:
                    list_vizhual_code_for_ticket.append(' ')

        list_numb_amount_box = []
        for i in sorted_dict_numb_barcode.keys():
            for k in dict_numb_amount_box.keys():
                if i == k:
                    list_numb_amount_box.append(int(dict_numb_amount_box[k]))

        list_numb_amount_product_in_box = []
        for i in sorted_dict_numb_barcode.keys():
            for k in dict_numb_amount_product_in_box.keys():
                if i == k:
                    list_numb_amount_product_in_box.append(
                        int(dict_numb_amount_product_in_box[k]))

        global amount_list_vizhual_code_for_ticket
        global amount_list_numb_amount_product_in_box
        global amount_sorted_dict_numb_barcode_values
        global amount_list_vizhual_code_for_people

        amount_list_vizhual_code_for_ticket = []
        for code in range(len(list_vizhual_code_for_ticket)):
                amount_list_vizhual_code_for_ticket.extend(
                    [str(list_vizhual_code_for_ticket[code])
                    for i in range(int(list_numb_amount_box[code]))])

        amount_list_numb_amount_product_in_box = []
        for code in range(len(list_numb_amount_product_in_box)):
                amount_list_numb_amount_product_in_box.extend(
                    [int(list_numb_amount_product_in_box[code])
                    for i in range(int(list_numb_amount_box[code]))])

        amount_sorted_dict_numb_barcode_values = []
        for code in range(len(sorted_dict_numb_barcode_values)):
                amount_sorted_dict_numb_barcode_values.extend(
                    [str(sorted_dict_numb_barcode_values[code])
                    for i in range(int(list_numb_amount_box[code]))])

        list_vizhual_code_for_people.sort()
        amount_list_vizhual_code_for_people = []
        for code in range(len(list_vizhual_code_for_people)):
                amount_list_vizhual_code_for_people.extend(
                    [str(list_vizhual_code_for_people[code])
                    for i in range(int(list_numb_amount_box[code]))])
            

    def create_xls_file(self):
        """Функция создает нужный excel файл"""
        # Создаем файл excel
        new_xls = openpyxl.Workbook()
        sheet = new_xls.active
        # Обозначаем название столбцов
        sheet['A1'] = 'Визуальный код'
        sheet['B1'] = 'баркод товара'
        sheet['C1'] = 'количество товара'
        sheet['D1'] = 'шк короба'
        sheet['E1'] = 'маркер'

        # Заполняем таблицу данными
        df = pd.DataFrame.from_dict({'Визуальный код': amount_list_vizhual_code_for_people,
                                     'баркод товара': amount_sorted_dict_numb_barcode_values,
                                     'количество товара': amount_list_numb_amount_product_in_box,
                                     'шк короба':barcode_boxes_list,
                                     'маркер': 1})
        new_xls.save(f'{file_name_dir_product}/raw_excel.xlsx')
        df.to_excel(f'{file_name_dir_product}/raw_excel.xlsx', header=True, index=False)

    def create_xls_file_for_WB(self):
        """Функция создает нужный excel файл"""
        # Создаем файл excel
        new_xls = openpyxl.Workbook()
        sheet = new_xls.active
        # Обозначаем название столбцов
        sheet['A1'] = 'баркод товара'
        sheet['B1'] = 'кол-во товаров'
        sheet['C1'] = 'шк короба'
        sheet['D1'] = 'срок годности'
        # Заполняем таблицу данными
        df = pd.DataFrame.from_dict({'баркод товара':amount_sorted_dict_numb_barcode_values,
                                     'кол-во товаров': amount_list_numb_amount_product_in_box,
                                     'шк короба':barcode_boxes_list,
                                     'срок годности': ' '})
        
        
        new_xls.save(f'{file_name_dir_product}/Файл для загрузки на ВБ.xlsx')
        df.to_excel(f'{file_name_dir_product}/Файл для загрузки на ВБ.xlsx', header=True, index=False)

        wb = openpyxl.load_workbook(f'{file_name_dir_product}/Файл для загрузки на ВБ.xlsx')
        al = Alignment()
        ws = wb.active
        
        ws['A1'].border = Border()
        ws['B1'].border = Border()
        ws['C1'].border = Border()
        ws['D1'].border = Border()

        ws['A1'].alignment = al
        ws['B1'].alignment = al
        ws['C1'].alignment = al
        ws['D1'].alignment = al

        ws['A1'].font = Font(bold=False)
        ws['B1'].font = Font(bold=False)
        ws['C1'].font = Font(bold=False)
        ws['D1'].font = Font(bold=False)

        wb.save(f'{file_name_dir_product}/Файл для загрузки на ВБ.xlsx')

    def style_xls_file(self):
        """Функция рисует границы ячейкам и заливает цветом"""
        # Читаем созданный файл
        wb = openpyxl.load_workbook(f'{file_name_dir_product}/raw_excel.xlsx')
        ws = wb.active
        # Задаем параметр выравнивания в ячейке - по центру
        al = Alignment(horizontal="center", vertical="center")
        # Задаем толщину и цвет обводки ячейки
        thick = Side(border_style="medium", color="000000")
        thin = Side(border_style="thin", color="000000")
        for i in range(1, len(amount_sorted_dict_numb_barcode_values)):
            for c in ws[f'A{i+2}:D{i+2}']:
                    c[0].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[1].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[2].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[3].border = Border(top=thin, right=thin, left=thin, bottom = thin)
        # Отмечаем какие ячейки нужно выделять цветом и обводить
        for i in range(1, len(amount_sorted_dict_numb_barcode_values)):
            if ws[f'B{i+2}'].value == ws[f'B{(i-1)+2}'].value:
                ws[f'E{i+2}'].value = ws[f'E{(i-1)+2}'].value    
            elif ws[f'B{i+2}'].value != ws[f'B{(i-1)+2}'].value and ws[f'E{(i-1)+2}'].value == 0:
                for c in ws[f'A{i+2}:D{i+2}']:
                    c[0].border = Border(top=thick, left=thick, bottom = thin, right = thin)
                    c[1].border = Border(top=thick, left=thin, bottom = thin, right = thin)
                    c[2].border = Border(top=thick, left=thin, bottom = thin, right = thin)
                    c[3].border = Border(top=thick, right=thick, left=thin, bottom = thin)
                ws[f'E{i+2}'].value = 1
            elif ws[f'B{i+2}'].value != ws[f'B{(i-1)+2}'].value and ws[f'E{(i-1)+2}'].value != 0:
                for c in ws[f'A{i+2}:D{i+2}']:
                    for j in range(4):
                        c[j].border = Border(top=thick, left=thin, bottom = thin, right = thin)
                ws[f'E{i+2}'].value = ws[f'E{i+2}'].value - 1
        # Обводими выделяем необходимые ячейки
        for i in range(len(amount_sorted_dict_numb_barcode_values)):
            if ws[f'E{i+2}'].value == 1:
                for c in ws[f'A{i+2}:D{i+2}']:
                    for j in range(4):
                        c[j].fill = PatternFill('solid', fgColor="FFFF00")
                        c[j].alignment = al

                for c in ws[f'A{i+2}:D{i+2}']:
                    c[0].border = Border(left=thick, top=thin, bottom = thin, right = thin)
                    c[3].border = Border(right=thick, top=thin, left=thin, bottom = thin)
            for c in ws[f'A{i+2}:D{i+2}']:
                for j in range(4):
                    c[j].alignment = al
            if ws[f'E{i+2}'].value == 1 and ws[f'E{i+1}'].value == 0:
                for c in ws[f'A{i+2}:D{i+2}']:
                    c[0].border = Border(top=thick, left=thick, bottom = thin, right = thin)
                    c[3].border = Border(top=thick, right=thick, left=thin, bottom = thin)
        # Задаем ширину ячеек, чтобы был виден весь текст
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 18

        # Удаляем колонку для маркировки и сохраняем файл
        ws.delete_cols(5)
        wb.save(f'{file_name_dir_product}/raw_excel.xlsx')


    def export_to_pdf(self):
        """Функция экспортирует xlsx в pdf"""
        xl = DispatchEx("Excel.Application")
        xl.DisplayAlerts = False
        wb = xl.Workbooks.Open(f'{file_name_dir_product}/raw_excel.xlsx')
        xl.CalculateFull()
        pythoncom.PumpWaitingMessages()
        wb.ExportAsFixedFormat(0, f'{file_name_dir_product}/расшифровка для сборки.pdf')
        wb.Close()

    def create_list_barcode(self):
        """Формирует список штрихкодов для печати"""
        # Читаем в файле excel нужные столбцы
        main_file_location = f'{file_name_dir_product}/raw_excel.xlsx'

        excel_data = pd.read_excel(main_file_location)

        data = pd.DataFrame(
            excel_data, columns=['баркод товара', 'шк короба'])
        barcode_box_list = data['шк короба'].to_list()

        # Подключаем шрифты чтобы писать текст
        font = ImageFont.truetype("arial.ttf", size=120)
        barcode_size = [img2pdf.mm_to_pt(70), img2pdf.mm_to_pt(49.5)]
        layout_function = img2pdf.get_layout_fun(barcode_size)

        list_for_pdfmerge = []

        for i in range(len(barcode_box_list)):
            render_options = {
                "module_width": 1,
                "module_height": 50,
                "font_size": 0,
                "text_distance": 8,
                "quiet_zone": 8
            }
            barcode = Code128(
                f'{barcode_box_list[i]}',
                writer=ImageWriter()
                ).render(render_options)
            main_image = Image.new('RGB', (1980, 1400), color=('#ffffff'))
            draw_text = ImageDraw.Draw(main_image)
            # Длина подложки
            w_width = round(main_image.width/2)
            # Длина штрихкода
            w = round(barcode.width/2)
            # Расположение штрихкода по центру
            position = w_width - w
            # Вставляем штрихкод в основной фон
            main_image.paste(barcode, ((w_width - w), 350))
            draw_text.text(
                (1520, 150),
                f'{amount_list_vizhual_code_for_people[i]}\n',
                font=font,
                fill=('#000000'), stroke_width=1
                )
            draw_text.text(
                (position+20, 150),
                f'{barcode_box_list[i]}\n',
                font=font,
                fill=('#000000'), stroke_width=1
                )
            draw_text.text(
                (position+100, 1000),
                f'{amount_list_vizhual_code_for_ticket[i]}\n',
                font=font,
                fill=('#000000'), stroke_width=1
                )
            main_image.save(f'cache_dir/{barcode_box_list[i]}.png')
            main_image.close()
            pdf = img2pdf.convert(
                f'cache_dir/{barcode_box_list[i]}.png',
                layout_fun=layout_function)
            with open(f'cache_dir/{barcode_box_list[i]}.pdf', 'wb') as f:
                f.write(pdf)
            list_for_pdfmerge.append(f'cache_dir/{barcode_box_list[i]}.pdf')
            sg.one_line_progress_meter('Создаю штрихкоды', i+1, len(barcode_box_list), 'Создаю штрихкоды')
        return list_for_pdfmerge


    def print_barcode_to_pdf(self):
        """Создает pdf файл для печати"""
        pdf_filenames = self.create_list_barcode()
        file_name = f'{file_name_dir_product}/box-barcode {time.strftime("%Y-%m-%d %H-%M")}.pdf'
        print_barcode_to_pdf2(pdf_filenames, file_name)

        # Очистка кеша

        dir = 'cache_dir/'
        filelist = glob.glob(os.path.join(dir, "*"))
        for f in filelist:
            try:
                os.remove(f)
            except Exception:
                print(' ')
        print('Done')


if __name__ == '__main__':
    import sys
    app = QApplication(sys.argv)
    w = BarcodeBoxPrintToPdf()
    w.show()
    sys.exit(app.exec())
