import pdfplumber
import glob
from collections import Counter
from pathlib import Path
import shutil
import time
from contextlib import closing
import io
import pythoncom
from win32com.client import DispatchEx
from draganddrop import DropMainWindow
from PyPDF3 import PdfFileReader
import os
import pandas as pd
import PySimpleGUI as sg
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, PatternFill, Side, Font

from helper_func import (print_barcode_to_pdf2,
                         qrcode_print_to_file_main, design_barcodes)



class CheckAllContent(DropMainWindow):
    """Первичный анализ всех входящих файлов"""
    def __init__(self):
        DropMainWindow.__init__(self)

    def pdf_file_analyze(self):
        """Разбивает файлы на списки по расширениям.
        Анализирует каждое расширение в отдельном цикле"""
        files_list = self.func_returne_list()
        pdf_files_list = []
        self.xls_files_list = []
        i = 1
        for file in files_list:
            path = Path(file)
            if str(os.path.basename(path).split('.')[1]) == 'pdf':
                pdf_files_list.append(file)
            else:
                self.xls_files_list.append(file)

        for file in pdf_files_list:
            path = Path(file)
            self.file_name_dir = path.parent
            i +=1
            sg.one_line_progress_meter('Проверяю файлы',
                                       i+1,
                                       len(pdf_files_list),
                                       'Проверяю файлы')
            with open(file, "rb") as filehandle:  
                pdf = PdfFileReader(filehandle)
                page1 = pdf.getPage(0) 
                if str(page1['/MediaBox']) == str([0, 0, 113.39, 85.04]):
                    self.wb_qrcode.setChecked(True)
                    continue

            with pdfplumber.open(file) as pdf:
                first_page = pdf.pages[0]
                if 'Лист подбора WB' in first_page.extract_text():
                    pdf.close()
                    self.wb_list.setChecked(True)
                    continue

            with open(file, "rb") as filehandle:  
                pdf = PdfFileReader(filehandle)
                page1 = pdf.getPage(0)
                if str(page1['/MediaBox']) == str([0, 0, 164.41, 113.39]):
                    self.wb_act.setChecked(True)
                    continue

            with pdfplumber.open(file) as pdf:
                first_page = pdf.pages[0]
                if 'ТРАНСПОРТНАЯ НАКЛАДНАЯ' in first_page.extract_text():
                    pdf.close()
                    self.ozon_act.setChecked(True)
                    continue

            if 'ticket-' in str(path):
                self.ozon_tickets.setChecked(True)
                continue

            elif 'act_TMU' in str(path):
                self.yandex_act.setChecked(True)
                continue
                
            elif 'order_service_first_mile' in str(path):
                self.yandex_task_pdf.setChecked(True)
                continue

            elif 'shipment_orders_labels' in str(path):
                self.yandex_tickets.setChecked(True)
                continue

        j = 1
        for file in self.xls_files_list:
            path = Path(file)
            j +=1
            sg.one_line_progress_meter('Проверяю файлы',
                                       j+1,
                                       len(self.xls_files_list),
                                       'Проверяю файлы')
            if 'postings' in str(path):
                self.ozon_task.setChecked(True)
            elif 'order_service_first' in str(path):
                self.yandex_task.setChecked(True)
            elif 'wb-gi' in str(path):
                self.wb_task.setChecked(True)


class RenamePdf(DropMainWindow):
    """Переименовывает PDF файлы и создает сводный файл для OZON"""
    def __init__(self):
        DropMainWindow.__init__(self)
        self.qrcode_file = 1
        self.name_pivot_xls = 1
        self.xls_files_list
        self.file_name_dir
        self.filename_for_qrcodes
        self.full_filename_add_qrcode
        self.filename_add_qrcodes

    def pdf_file_analyze(self):
        files_list = self.func_returne_list()
        pdf_files_list = []
        self.xls_files_list = []
        for file in files_list:
            path = Path(file)
            if str(os.path.basename(path).split('.')[1]) == 'pdf':
                pdf_files_list.append(file)
            else:
                self.xls_files_list.append(file)

        for file in pdf_files_list:
            path = Path(file)
            self.file_name_dir = path.parent
            with open(file, "rb") as filehandle:  
                pdf = PdfFileReader(filehandle) 
                page1 = pdf.getPage(0) 
                if str(page1['/MediaBox']) == str([0, 0, 113.39, 85.04]):
                    self.filename_for_qrcodes = os.path.basename(path).split('.')[0]
                    self.qrcode_file = file
                    continue
            
            with pdfplumber.open(file) as pdf:
                first_page = pdf.pages[0]
                if 'Лист подбора WB' in first_page.extract_text():
                    pdf.close()
                    new_name = os.path.join(self.file_name_dir, 'WB-ООО лист подбора.pdf')
                    os.rename(file, new_name)
                    continue

            with open(file, "rb") as filehandle:  
                pdf = PdfFileReader(filehandle)
                page1 = pdf.getPage(0)
                if str(page1['/MediaBox']) == str([0, 0, 164.41, 113.39]):
                    self.full_filename_add_qrcode = file
                    self.filename_add_qrcodes = os.path.basename(path).split('.')[0]
                    if not self.checkbox_add_qrcode.isChecked():
                        filehandle.close()
                        new_name = os.path.join(self.file_name_dir, 'WB-ООО акты.pdf')
                        os.rename(file, new_name)
                    else:
                        None
                    continue

            with pdfplumber.open(file) as pdf:
                first_page = pdf.pages[0]
                if 'ТРАНСПОРТНАЯ НАКЛАДНАЯ' in first_page.extract_text():
                    pdf.close()
                    new_name = os.path.join(self.file_name_dir, 'OZON-ООО акты.pdf')
                    os.rename(file, new_name)
                    continue

            if 'ticket-' in str(path):
                new_name = os.path.join(self.file_name_dir, 'OZON-ООО этикетки.pdf')
                os.rename(file, new_name)
                continue
            elif 'act_TMU' in str(path):
                new_name = os.path.join(self.file_name_dir, 'YANDEX-ООО акты.pdf')
                os.rename(file, new_name)
                continue
            elif 'order_service_first_mile' in str(path):
                new_name = os.path.join(self.file_name_dir, 'YANDEX-ООО лист подбора.pdf')
                os.rename(file, new_name)
                continue
            elif 'shipment_orders_labels' in str(path):
                new_name = os.path.join(self.file_name_dir, 'YANDEX-ООО этикетки.pdf')
                os.rename(file, new_name)
                continue

    def create_ozone_selection_sheet_pdf(self):
        for file in self.xls_files_list:
            path = Path(file)
            file_name_dir = path.parent
            if 'postings' in str(path):          
                
                excel_data = pd.read_csv(file, delimiter=';')

                number_of_departure_oz = excel_data['Номер отправления'].to_list()
                product_name_oz = excel_data['Наименование товара'].to_list()
                name_for_print_oz = excel_data['Артикул'].to_list()
                amount_for_print_oz = excel_data['Количество'].to_list()

                ozone_selection_sheet_xls = openpyxl.Workbook()
                create = ozone_selection_sheet_xls.create_sheet(title = 'pivot_list', index = 0)
                sheet = ozone_selection_sheet_xls['pivot_list']
                sheet['A1'] = 'Номер отправления'
                sheet['B1'] = 'Наименование товара'
                sheet['C1'] = 'Артикул'
                sheet['D1'] = 'Количество'

                al = Alignment(horizontal="center", vertical="center", wrap_text=True)
                al2 = Alignment(vertical="center", wrap_text=True)
                thin = Side(border_style="thin", color="000000")
                thick = Side(border_style="medium", color="000000")
                pattern = PatternFill('solid', fgColor="fcff52")
                new_dict = {}
                for i in range(len(number_of_departure_oz)):
                    new_dict[i] = [number_of_departure_oz[i],
                                   product_name_oz[i], name_for_print_oz[i],
                                   amount_for_print_oz[i]]

                sorted_dict = dict(sorted(new_dict.items(), key=lambda item: item[1][0][-6:]))

                upd_number_of_departure_oz = []
                upd_product_name_oz = []
                upd_name_for_print_oz = []
                upd_amount_for_print_oz = []

                for key, value in sorted_dict.items():
                    upd_number_of_departure_oz.append(value[0])
                    upd_product_name_oz.append(value[1])
                    upd_name_for_print_oz.append(value[2])
                    upd_amount_for_print_oz.append(value[3])

                for i in range(len(upd_number_of_departure_oz)):
                    create.cell(row=i+2,column=1).value = upd_number_of_departure_oz[i]
                    create.cell(row=i+2,column=2).value = upd_product_name_oz[i]
                    create.cell(row=i+2,column=3).value = upd_name_for_print_oz[i]
                    create.cell(row=i+2,column=4).value = upd_amount_for_print_oz[i]
                for i in range(1, len(upd_number_of_departure_oz)+2):
                    for c in create[f'A{i}:D{i}']:
                        c[0].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                        c[1].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                        c[2].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                        c[3].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                        c[3].alignment = al
                        for j in range(3):
                            c[j].alignment = al2
                sg.one_line_progress_meter('Создаю  xls файл OZON',
                                       i+1,
                                       len(self.xls_files_list),
                                       'Создаю  xls файл OZON')

                create.column_dimensions['A'].width = 18
                create.column_dimensions['B'].width = 38
                create.column_dimensions['C'].width = 18
                create.column_dimensions['D'].width = 10

                name_for_file = f'{file_name_dir}/OZON-ООО лист подбора'

                ozone_selection_sheet_xls.save(f'{name_for_file}.xlsx')
                w_b2 = load_workbook(f'{name_for_file}.xlsx')
                source_page2 = w_b2.active
                number_of_departure_oz = source_page2['A']
                amount_for_print_oz = source_page2['D']

                for i in range(1, len(number_of_departure_oz)+2):
                    if i < len(number_of_departure_oz)-1:
                        for c in source_page2[f'A{i+1}:D{i+1}']:
                            if (number_of_departure_oz[i].value == number_of_departure_oz[i+1].value
                                and number_of_departure_oz[i].value != number_of_departure_oz[i-1].value):
                                c[0].border = Border(top=thick, left=thick, bottom = thin, right = thin)
                                c[1].border = Border(top=thick, left=thin, bottom = thin, right = thin)
                                c[2].border = Border(top=thick, left=thin, bottom = thin, right = thin)
                                c[3].border = Border(top=thick, left=thin, bottom = thin, right = thick)
                                for j in range(4):
                                    c[j].fill = pattern
                            if (number_of_departure_oz[i].value == number_of_departure_oz[i+1].value
                                and number_of_departure_oz[i].value == number_of_departure_oz[i-1].value):
                                c[0].border = Border(top=thin, left=thick, bottom = thin, right = thin)
                                c[1].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                                c[2].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                                c[3].border = Border(top=thin, left=thin, bottom = thin, right = thick)
                                for j in range(4):
                                    c[j].fill = pattern
                            elif (number_of_departure_oz[i].value != number_of_departure_oz[i+1].value
                                  and number_of_departure_oz[i].value == number_of_departure_oz[i-1].value):
                                c[0].border = Border(top=thin, left=thick, bottom = thick, right = thin)
                                c[1].border = Border(top=thin, left=thin, bottom = thick, right = thin)
                                c[2].border = Border(top=thin, left=thin, bottom = thick, right = thin)
                                c[3].border = Border(top=thin, left=thin, bottom = thick, right = thick)
                                for j in range(4):
                                    c[j].fill = pattern
                            if amount_for_print_oz[i].value > 1:
                                c[0].border = Border(top=thick, left=thick, bottom = thick, right = thin)
                                c[1].border = Border(top=thick, left=thin, bottom = thick, right = thin)
                                c[2].border = Border(top=thick, left=thin, bottom = thick, right = thin)
                                c[3].border = Border(top=thick, left=thin, bottom = thick, right = thick)
                                for j in range(4):
                                    c[j].fill = pattern
                            sg.one_line_progress_meter('Создаю  xls файл OZON',
                                       i+1,
                                       len(number_of_departure_oz),
                                       'Создаю  xls файл OZON')

                w_b2.save(f'{name_for_file}.xlsx')

                xl = DispatchEx("Excel.Application")
                xl.DisplayAlerts = False
                wb = xl.Workbooks.Open(f'{name_for_file}.xlsx')
                xl.CalculateFull()
                pythoncom.PumpWaitingMessages()
                try:
                    wb.ExportAsFixedFormat(0, f'{name_for_file}.pdf')
                except Exception as e:
                    print("Failed to convert in PDF format.Please confirm environment meets all the requirements  and try again")
                finally:
                    wb.Close()
                if os.path.isfile(f'{name_for_file}.xlsx'): 
                    os.remove(f'{name_for_file}.xlsx') 
                    print("success") 
                else: 
                    print("File doesn't exists!")
                break


class CreatePivoteExcel(RenamePdf):
    def __init__(self):
        RenamePdf.__init__(self)
        self.list_with_articles
        self.name_article_wb
        self.amount_article_wb
        self.name_article_wb
        
    def xls_file_analyze(self):
        self.list_with_articles = []
        self.name_article_wb = []
        self.amount_article_wb = []
        self.name_article_oz = []
        self.amount_article_oz = []
        self.name_article_ya = []
        self.amount_article_ya = []
        for file in self.xls_files_list:
            path = Path(file)
            file_name_dir = path.parent

            if 'wb-' in str(path):
                raw_name_article_wb = []
                excel_data = pd.read_excel(file)
                data = pd.DataFrame(
                    excel_data, columns=['Артикул продавца', 'Количество'])
                name_for_print_wb = data['Артикул продавца'].to_list()
                for article_wb in name_for_print_wb:
                    raw_name_article_wb.append(article_wb)
                    self.list_with_articles.append(article_wb)
                raw_dict_name_count = Counter(raw_name_article_wb)
                for art, amo in raw_dict_name_count.items():
                    self.name_article_wb.append(art)
                    self.amount_article_wb.append(amo)
                continue

            elif 'order_service_first_mile' in str(path):
                raw_name_article_ya = []
                excel_data = pd.read_excel(file, skiprows=1)
                data = pd.DataFrame(
                    excel_data, columns=['Ваш SKU'])
                name_for_print_ya = data['Ваш SKU'].to_list()
                for article_ya in name_for_print_ya:
                    raw_name_article_ya.append(article_ya)
                    self.list_with_articles.append(article_ya)
                raw_dict_name_count_ya = Counter(raw_name_article_ya)
                for art_ya, amo_ya in raw_dict_name_count_ya.items():
                    self.name_article_ya.append(art_ya)
                    self.amount_article_ya.append(amo_ya)
                continue

            elif 'postings' in str(path):
                raw_name_article_oz = []
                excel_data = pd.read_csv(file, delimiter=';')
                name_for_print_oz = excel_data['Артикул'].to_list()
                for article_oz in name_for_print_oz:
                    raw_name_article_oz.append(article_oz)
                    self.list_with_articles.append(article_oz)
                raw_dict_name_count_oz = Counter(raw_name_article_oz)
                for art_oz, amo_oz in raw_dict_name_count_oz.items():
                    self.name_article_oz.append(art_oz)
                    self.amount_article_oz.append(amo_oz)
                continue

            sg.one_line_progress_meter('Создаю общий xls файл',
                                       file+1,
                                       len(self.xls_files_list),
                                       'Создаю общий xls файл')
    def create_pivot_xls(self):

        unsort_data = Counter(self.list_with_articles)
        sorted_data_for_pivot_xls = dict(sorted(unsort_data.items(), key=lambda v: v[0].upper()))

        pivot_xls = openpyxl.Workbook()
        create = pivot_xls.create_sheet(title = 'pivot_list', index = 0)
        sheet = pivot_xls['pivot_list']
        sheet['A1'] = 'Артикул продавца'
        sheet['B1'] = 'На производство'
        sheet['C1'] = 'Всего для FBS'
        sheet['D1'] = 'FBS WB'
        sheet['E1'] = 'FBS Ozon'
        sheet['F1'] = 'FBS Yandex'

        COUNT_HELPER = 2
        for key, value in sorted_data_for_pivot_xls.items():
            create.cell(row=COUNT_HELPER,column=1).value=key
            create.cell(row=COUNT_HELPER,column=3).value=value
            COUNT_HELPER += 1
        self.name_pivot_xls = f'{self.file_name_dir}/На производство.xlsx'
        pivot_xls.save(self.name_pivot_xls)

        w_b = load_workbook(self.name_pivot_xls)
        source_page = w_b.active
        name_article = source_page['A']

        for i in range(1, len(name_article)):
            for j in range(len(self.name_article_oz)):
                if name_article[i].value == self.name_article_oz[j]:
                    source_page.cell(row=i+1,column=5).value = self.amount_article_oz[j]

            for k in range(len(self.name_article_ya)):
                if name_article[i].value == self.name_article_ya[k]:
                    source_page.cell(row=i+1,column=6).value = self.amount_article_ya[k]
            
            for m in range(len(self.name_article_wb)):
                if name_article[i].value == self.name_article_wb[m]:
                    source_page.cell(row=i+1,column=4).value = self.amount_article_wb[m]
        w_b.save(self.name_pivot_xls)

        w_b2 = load_workbook(self.name_pivot_xls)
        source_page2 = w_b2.active
        amount_all_fbs = source_page2['C']
        amount_for_production = source_page2['B']

        PROD_DETAIL_CONST = 4
        for r in range(1, len(amount_all_fbs)):
            # Заполняет столбец ['B'] = 'Производство'
            if amount_all_fbs[r].value == 1:
                source_page2.cell(row=r+1,column=2).value = int(PROD_DETAIL_CONST)
            elif 2<= int(amount_all_fbs[r].value) <= PROD_DETAIL_CONST-1:
                source_page2.cell(row=r+1,column=2).value= int(2 * PROD_DETAIL_CONST)
            elif PROD_DETAIL_CONST<= int(amount_all_fbs[r].value) <= 2 * PROD_DETAIL_CONST - 1:
                source_page2.cell(row=r+1,column=2).value= int(3 * PROD_DETAIL_CONST)
            else:
                source_page2.cell(row=r+1,column=2).value= ' '
        w_b2.save(self.name_pivot_xls)

        w_b2 = load_workbook(self.name_pivot_xls)
        source_page2 = w_b2.active
        amount_all_fbs = source_page2['C']
        al = Alignment(horizontal="center", vertical="center")
        # Задаем толщину и цвет обводки ячейки
        font_bold = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="medium", color="000000")
        for i in range(len(amount_all_fbs)):
            for c in source_page2[f'A{i+1}:F{i+1}']:
                if i == 0:
                    c[0].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[1].border = Border(top=thick, left=thick, bottom = thin, right = thick)
                    c[2].border = Border(top=thick, left=thick, bottom = thin, right = thick)
                    c[3].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[4].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[5].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                elif i == len(amount_all_fbs)-1:
                    c[0].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[1].border = Border(top=thin, left=thick, bottom = thick, right = thick)
                    c[2].border = Border(top=thin, left=thick, bottom = thick, right = thick)
                    c[3].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[4].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[5].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                else:
                    c[0].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[1].border = Border(top=thin, left=thick, bottom = thin, right = thick)
                    c[2].border = Border(top=thin, left=thick, bottom = thin, right = thick)
                    c[3].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[4].border = Border(top=thin, left=thin, bottom = thin, right = thin)
                    c[5].border = Border(top=thin, left=thin, bottom = thin, right = thin)

                c[0].alignment = al
                c[1].alignment = al
                c[2].alignment = al
                c[3].alignment = al
                c[4].alignment = al
                c[5].alignment = al
                sg.one_line_progress_meter('Навожу красоту в xls файле',
                                       i+1,
                                       len(amount_all_fbs),
                                       'Навожу красоту в xls файле')
        source_page2.column_dimensions['A'].width = 18
        source_page2.column_dimensions['B'].width = 18
        source_page2.column_dimensions['C'].width = 15
        source_page2.column_dimensions['D'].width = 10
        source_page2.column_dimensions['E'].width = 10
        source_page2.column_dimensions['F'].width = 12
        w_b2.save(self.name_pivot_xls)


class PrintTicket(DropMainWindow):
    def __init__(self):
        super().__init__(self)
        self.qrcode_file = RenamePdf.qrcode_file
        self.name_pivot_xls = RenamePdf.name_pivot_xls
        self.name_article_wb = CreatePivoteExcel.name_article_wb
        self.amount_article_wb = CreatePivoteExcel.amount_article_wb
        self.file_name_dir = RenamePdf.file_name_dir
        self.filename_for_qrcodes = RenamePdf.filename_for_qrcodes
        self.full_filename_add_qrcode = RenamePdf.full_filename_add_qrcode
        self.filename_add_qrcodes = RenamePdf.filename_add_qrcodes

    def qrcode_print_to_file(self):
        """Создает QR коды в необходимом формате"""
        global pdf_filenames_qrcode
        pdf_filenames_qrcode = qrcode_print_to_file_main(
            self.qrcode_file, self.filename_for_qrcodes)
        if self.checkbox_add_qrcode.isChecked:
            list_add_qrcode = qrcode_print_to_file_main(
            self.full_filename_add_qrcode, self.filename_add_qrcodes)
            amount_qrcodes = int(sum(self.amount_article_wb)) // 20
            if int(sum(self.amount_article_wb)) % 20 > 0:
                amount_qrcodes = amount_qrcodes + 1
            for i in list_add_qrcode:
                for j in range(amount_qrcodes):
                    pdf_filenames_qrcode.append(i)

    def stream_dropbox_file(self, path):
        _,res=self.dbx.files_download(path)
        with closing(res) as result:
            byte_data=result.content
            return io.BytesIO(byte_data)

    def create_list_barcode(self):
        """Формирует список штрихкодов для печати"""
        address, self.main_file1 = self.showDetails()
        main_file_location = self.main_file1
        png_file = 'programm_data/eac.png'
        excel_data = pd.read_excel(main_file_location)
        data = pd.DataFrame(
            excel_data, columns=['Артикул продавца', 'Баркод товара', 'Наименование'])
        lenght_list = len(data['Артикул продавца'].to_list())

        first_file = pd.read_excel(self.name_pivot_xls)
        first_file_data = pd.DataFrame(first_file,
                                       columns=['Артикул продавца',
                                                'На производство',
                                                'Всего для FBS'])

        name_ticket_for_print = first_file_data['Артикул продавца'].to_list()

        self.name_article_wb 
        self.amount_article_wb
        barcodelist_list = data['Баркод товара'].to_list()
        article_list = data['Артикул продавца'].to_list()
        article_list = [item.capitalize() for item in article_list]
        self.name_article_wb = [item.capitalize() for item in self.name_article_wb]
        name_ticket_for_print = [item.capitalize() for item in name_ticket_for_print]

        name_list = data['Наименование'].to_list()

        address, self.main_file1 = self.showDetails()
        excel_data2 = openpyxl.load_workbook(address)
        sheet = excel_data2.active

        design_barcodes(name_ticket_for_print, article_list,
                        self.name_pivot_xls, main_file_location,
                        barcodelist_list, name_list, sheet)

        pdf_filenames = glob.glob('cache_dir/*.pdf')
        self.new_list = []
        for i in range(len(self.name_article_wb)):
            for j in pdf_filenames:
                if str(self.name_article_wb[i]) == str(Path(j).stem):
                    while self.amount_article_wb[i] > 0:
                        self.new_list.append(j)
                        self.amount_article_wb[i] -= 1

        
        self.new_list.sort()
        for j in pdf_filenames_qrcode:
            self.new_list.append(j)
        return self.new_list
    
    def create_pdf_file_ticket_for_complect(self):
        """Создает PDF-файл для формирования остатки FBS"""
        first_file = pd.read_excel(self.name_pivot_xls)
        
        first_file_data = pd.DataFrame(first_file,
                                       columns=['Артикул продавца',
                                                'На производство',
                                                'Всего для FBS'])
        
        name_ticket_for_print = first_file_data['Артикул продавца'].to_list()
        amount_for_production = first_file_data['На производство'].to_list()
        amount_for_fbs = first_file_data['Всего для FBS'].to_list()
    
        amount_for_print = []
        for i in range(len(name_ticket_for_print)):
            if str(amount_for_production[i]) == ' ':
                amount_for_print.append(1)
            else:
                quantity_for_print = int(amount_for_production[i]) - int(amount_for_fbs[i])
                amount_for_print.append(quantity_for_print)

        pdf_filenames = glob.glob('cache_dir/*.pdf')
        list_pdf_file_ticket_for_complect = []
        for i in range(len(name_ticket_for_print)):
            for j in pdf_filenames:
                if str(name_ticket_for_print[i]) == str(Path(j).stem):
                    while amount_for_print[i] > 0:
                        list_pdf_file_ticket_for_complect.append(j)
                        amount_for_print[i] -= 1
        file_name = (f'{self.file_name_dir}/Наклейки для комплектовщиков '
                         f'{time.strftime("%Y-%m-%d %H-%M")}.pdf')
        print_barcode_to_pdf2(list_pdf_file_ticket_for_complect, file_name)

    def print_barcode_in_pdf(self):
        """Создает pdf файл для печати"""
        file_name = (f'{self.file_name_dir}/WB-ООО этикетки FBS '
                         f'{time.strftime("%Y-%m-%d %H-%M")}.pdf')
        print_barcode_to_pdf2(self.new_list, file_name)
        
        # Очистка переменных перед новым запуском
        self.new_list.clear()
        
        # Очистка кеша
        dir = 'cache_dir/'
        filelist = glob.glob(os.path.join(dir, "*"))
        for f in filelist:
            try:
                os.remove(f)
            except Exception:
                print('')
        
        dir = 'cache_dir_2/'
        filelist = glob.glob(os.path.join(dir, "*"))
        for f in filelist:
            try:
                os.remove(f)
            except Exception:
                print('')
        
        dir = 'cache_dir_3/'
        filelist = glob.glob(os.path.join(dir, "*"))
        for f in filelist:
            try:
                os.remove(f)
            except Exception:
                print('')
        print('Done')

        folder = 'cache_dir_3/'
        for filename in glob.glob(os.path.join(folder, "*")):
            file_path = os.path.join(folder, filename)
            try:
                if os.path.isfile(filename) or os.path.islink(filename):
                    os.unlink(filename)
                elif os.path.isdir(filename):
                    shutil.rmtree(filename)
            except Exception as e:
                print('')
        
    
